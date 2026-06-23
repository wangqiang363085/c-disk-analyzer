# -*- coding: utf-8 -*-
"""
C 盘空间分析工具
=================
- 扫描 C 盘各目录大小，找出空间占用大户
- 检测 AppData 缓存（剪映、美图、WPS 等）
- 生成带饼图的 Excel 报告 + 清理建议
- 可选一键清理 npm / pip 缓存

运行方式:
  python c_disk_analyzer.py
  或双击 dist/ 下的 exe
"""

import os
import sys
import subprocess
import threading
from datetime import datetime


def is_admin():
    try:
        import ctypes
        return ctypes.windll.shell32.IsUserAnAdmin()
    except:
        return False


def fmt(b):
    if b is None or b < 0:
        return "N/A"
    if b >= 1e9:
        return f"{b/1e9:.2f} GB"
    if b >= 1e6:
        return f"{b/1e6:.2f} MB"
    if b >= 1e3:
        return f"{b/1e3:.2f} KB"
    return f"{int(b)} B"


def dir_size(path, timeout=30, depth=0):
    if not os.path.exists(path):
        return None
    res = [None]

    def work():
        try:
            total = 0
            base = path.rstrip("\\").count("\\")
            for root, dirs, files in os.walk(path):
                if depth > 0 and root.count("\\") - base > depth:
                    dirs.clear()
                    continue
                for f in files:
                    try:
                        fp = os.path.join(root, f)
                        if os.path.exists(fp):
                            total += os.path.getsize(fp)
                    except Exception:
                        pass
            res[0] = total
        except Exception:
            pass

    t = threading.Thread(target=work, daemon=True)
    t.start()
    t.join(timeout)
    return res[0]


def scan():
    data = {"total": 0, "used": 0, "free": 0, "folders": [], "appdata": [], "caches": []}

    # 磁盘容量
    try:
        import ctypes
        fb = ctypes.c_ulonglong(0)
        tb = ctypes.c_ulonglong(0)
        ctypes.windll.kernel32.GetDiskFreeSpaceExW(
            "C:\\", None, ctypes.byref(tb), ctypes.byref(fb)
        )
        data["total"] = tb.value
        data["free"] = fb.value
        data["used"] = data["total"] - data["free"]
    except Exception:
        pass

    print(f"\n总容量: {fmt(data['total'])}")
    print(f"已用:   {fmt(data['used'])} ({data['used']/data['total']*100:.1f}%)")
    print(f"可用:   {fmt(data['free'])}")

    # 一级目录
    print("\n扫描一级目录...")
    skip = {
        "System Volume Information", "Recovery", "$Recycle.Bin",
        "Windows.old", "Config.Msi", "Documents and Settings"
    }
    try:
        for item in os.scandir("C:\\"):
            if not item.is_dir(follow_symlinks=False):
                continue
            if item.name in skip:
                continue
            sz = dir_size(item.path, timeout=35)
            if sz and sz > 0:
                data["folders"].append({"name": item.name, "size": sz})
    except Exception:
        pass
    data["folders"].sort(key=lambda x: x["size"], reverse=True)

    # AppData
    print("\n扫描 AppData 缓存...")
    local = os.environ.get("LOCALAPPDATA", "")
    if local and os.path.exists(local):
        for item in os.scandir(local):
            if item.is_dir():
                sz = dir_size(item.path, timeout=12, depth=3)
                if sz and sz > 100e6:
                    data["appdata"].append({"name": item.name, "size": sz})
    data["appdata"].sort(key=lambda x: x["size"], reverse=True)

    # 缓存目录
    print("\n扫描缓存目录...")
    cache_paths = [
        ("用户临时文件", os.environ.get("TEMP", "")),
        ("Windows 临时文件", os.environ.get("WINDIR", "") + "\\Temp"),
        ("Windows 更新缓存", os.environ.get("WINDIR", "") + "\\SoftwareDistribution\\Download"),
        ("npm 缓存", os.environ.get("APPDATA", "") + "\\npm-cache"),
        ("pip 缓存", os.environ.get("LOCALAPPDATA", "") + "\\pip\\cache"),
        ("浏览器缓存", os.environ.get("LOCALAPPDATA", "") + "\\Microsoft\\Windows\\INetCache"),
    ]
    for label, path in cache_paths:
        if os.path.exists(path):
            sz = dir_size(path, timeout=10, depth=2)
            if sz and sz > 1024 * 1024:
                data["caches"].append({"label": label, "size": sz})
    data["caches"].sort(key=lambda x: x["size"], reverse=True)

    return data


def gen_excel(data, path):
    # 延迟导入，避免没装 openpyxl 时直接报错
    import openpyxl
    from openpyxl.chart import PieChart, Reference
    from openpyxl.chart.label import DataLabelList
    from openpyxl.chart.series import DataPoint
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = openpyxl.Workbook()
    hf = Font(name="微软雅黑", bold=True, size=11, color="FFFFFF")
    hfill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    df = Font(name="微软雅黑", size=10)
    tf = Font(name="微软雅黑", bold=True, size=14, color="2F5496")
    stf = Font(name="微软雅黑", bold=True, size=12, color="C00000")
    rf = Font(name="微软雅黑", size=10, bold=True, color="FF0000")
    gf = Font(name="微软雅黑", size=10, bold=True, color="00B050")
    bd = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    gfill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    rfill = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")
    ofill = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
    yfill = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")

    def shdr(ws, hd, wd):
        for i, (h, w) in enumerate(zip(hd, wd), 1):
            c = ws.cell(row=1, column=i, value=h)
            c.font = hf; c.fill = hfill
            c.alignment = Alignment(horizontal='center', vertical='center')
            c.border = bd
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    def wr(ws, r, data, dc=None):
        for ci, v in enumerate(data, 1):
            c = ws.cell(row=r, column=ci, value=v)
            c.font = df
            c.alignment = Alignment(vertical='center', wrap_text=True)
            c.border = bd
            if ci == dc:
                c.alignment = Alignment(horizontal='center', vertical='center')
                if v == "能":
                    c.fill = gfill
                elif v == "不能":
                    c.fill = rfill
                elif v == "部分能":
                    c.fill = ofill
                else:
                    c.fill = yfill

    # === Sheet 1: 空间总结 + 饼图 + 建议（放在最前面）===
    ws_s = wb.active
    ws_s.title = "空间总结"
    fg = data["free"] / 1e9
    ug = data["used"] / 1e9
    tg = data["total"] / 1e9
    ws = sum(f["size"] for f in data["folders"] if "windows" in f["name"].lower()) / 1e9
    p64 = sum(f["size"] for f in data["folders"] if f["name"] == "Program Files") / 1e9
    p86 = sum(f["size"] for f in data["folders"] if f["name"] == "Program Files (x86)") / 1e9
    pd = sum(f["size"] for f in data["folders"] if f["name"] == "ProgramData") / 1e9
    at = sum(a["size"] for a in data["appdata"]) / 1e9

    cats = []
    if ws > 0:
        cats.append(("Windows系统", round(ws, 1)))
    if at > 0:
        cats.append(("AppData缓存", round(at, 1)))
    if p64 > 0:
        cats.append(("Program Files(64)", round(p64, 1)))
    if p86 > 0:
        cats.append(("Program Files(32)", round(p86, 1)))
    if pd > 0:
        cats.append(("ProgramData", round(pd, 1)))
    oth = max(0, ug - ws - p64 - p86 - pd - at)
    if oth > 1:
        cats.append(("其他系统文件", round(oth, 1)))
    cats.append(("可用空间", round(fg, 1)))

    for ri, row in enumerate([["分类", "大小(GB)"]] + [[n, s] for n, s in cats], 1):
        for ci, v in enumerate(row, 1):
            c = ws_s.cell(row=ri, column=ci, value=v)
            if ri == 1:
                c.font = hf; c.fill = hfill
                c.alignment = Alignment(horizontal='center', vertical='center')
                c.border = bd
            else:
                c.font = df; c.border = bd
    ws_s.column_dimensions['A'].width = 28
    ws_s.column_dimensions['B'].width = 15

    # 饼图
    pie = PieChart()
    pie.title = f"C盘{tg:.0f}GB空间分布"
    pie.style = 10
    pie.width = 22
    pie.height = 15
    labels = Reference(ws_s, min_col=1, min_row=2, max_row=len(cats) + 1)
    dr = Reference(ws_s, min_col=2, min_row=1, max_row=len(cats) + 1)
    pie.add_data(dr, titles_from_data=True)
    pie.set_categories(labels)
    pie.dataLabels = DataLabelList()
    pie.dataLabels.showPercent = True
    pie.dataLabels.showCatName = True
    pie.dataLabels.showLeaderLines = True
    for i, c in enumerate([
        "4472C4", "5B9BD5", "ED7D31", "A5A5A5", "FFC000",
        "FF6999", "264478", "70AD47", "BFBFBF"
    ]):
        pt = DataPoint(idx=i)
        pt.graphicalProperties.solidFill = c
        pie.series[0].data_points.append(pt)
    ws_s.add_chart(pie, "D1")

    # === Sheet 2: C盘目录总览 ===
    ws_d = wb.create_sheet("C盘目录总览")
    shdr(ws_d, ["目录路径", "说明", "大小", "能否删除", "备注"], [50, 22, 14, 12, 55])
    rows_d = []
    for f in data["folders"]:
        n = f["name"].lower()
        if "windows" in n:
            cd, nt = "不能", "系统核心不要动"
        elif "program files" in n:
            cd, nt = "不能", "通过设置卸载勿手动删"
        elif "programdata" in n:
            cd, nt = "不能", "删了导致软件异常"
        elif n == "users":
            cd, nt = "视情况", "桌面/文档是个人文件"
        elif n in ("tmp", "temp"):
            cd, nt = "能", "临时文件安全可删"
        else:
            cd, nt = "视情况", "自行判断"
        rows_d.append([f["name"], f["name"], fmt(f["size"]), cd, nt])
    for i, r in enumerate(rows_d, 2):
        wr(ws_d, i, r, dc=4)
    ws_d.auto_filter.ref = f"A1:E{len(rows_d)+1}"
    ws_d.freeze_panes = "A2"

    # === Sheet 3: 清理建议（按可释放空间大小排序）===
    ws_c = wb.create_sheet("清理建议")
    ws_c.column_dimensions['A'].width = 32
    ws_c.column_dimensions['B'].width = 20
    ws_c.column_dimensions['C'].width = 50

    row = 1
    ws_c.cell(row=row, column=1, value="清理建议").font = tf
    ws_c.merge_cells("A1:C1")

    # 一、能马上清理的 - 用 data["caches"] 实际扫描数据
    row = 3
    ws_c.cell(row=row, column=1, value="一、能马上清理的（安全）").font = stf
    headers_safe = ["项目", "可释放空间", "操作方法"]
    for i, h in enumerate(headers_safe):
        c = ws_c.cell(row=row+1, column=i+1, value=h)
        c.font = Font(name="微软雅黑", bold=True, size=10)
        c.border = bd

    row_data = row + 2
    total_safe = 0
    for cache in data["caches"]:
        ws_c.cell(row=row_data, column=1, value=f"  {cache['label']}").font = df
        ws_c.cell(row=row_data, column=2, value=fmt(cache['size'])).font = df
        ws_c.cell(row=row_data, column=2).alignment = Alignment(horizontal='right')
        ws_c.cell(row=row_data, column=3, value="cleanmgr 或手动删除").font = df
        for ci in range(1, 4):
            ws_c.cell(row=row_data, column=ci).border = bd
        total_safe += cache["size"]
        row_data += 1

    # 小计
    ws_c.cell(row=row_data, column=1, value="  合计").font = Font(name="微软雅黑", size=10, bold=True)
    ws_c.cell(row=row_data, column=2, value=fmt(total_safe)).font = Font(name="微软雅黑", size=10, bold=True, color="00B050")
    for ci in range(1, 4):
        ws_c.cell(row=row_data, column=ci).border = bd

    # 二、AppData 大缓存（需在软件内清理）
    row = row_data + 2
    ws_c.cell(row=row, column=1, value="二、需在软件内清理的缓存（按大小排序）").font = stf
    headers_app = ["软件", "缓存大小", "清理方法"]
    for i, h in enumerate(headers_app):
        c = ws_c.cell(row=row+1, column=i+1, value=h)
        c.font = Font(name="微软雅黑", bold=True, size=10)
        c.border = bd

    known_methods = {
        "jianyingpro": "剪映 > 设置 > 清理缓存",
        "meituapp": "美图 > 设置 > 清理缓存",
        "meitu": "美图 > 设置 > 清理缓存",
        "wechat": "工具 > 设置 > 清理缓存",
        "kingsoft": "WPS > 设置 > 清理缓存",
        "doubao": "豆包 > 设置 > 清理缓存",
        "jetbrains": "IDE: File > Invalidate Caches",
        "feishu": "飞书 > 设置 > 清理缓存",
        "google": "Chrome > 设置 > 清除数据",
        "qianniu": "千牛 > 设置 > 清理缓存",
        "dingtalk": "钉钉 > 设置 > 清理缓存",
        "openai": "工具 > 设置 > 清理缓存",
        "mozilla": "Firefox > 设置 > 清除数据",
        "microsoft": "系统组件，不要动",
        "pip": "pip cache purge",
        "npm": "npm cache clean --force",
    }

    r = row + 2
    total_app = 0
    for a in data["appdata"]:
        kl = a["name"].lower()
        method = "软件设置里清理"
        for key, m in known_methods.items():
            if key in kl:
                method = m
                break
        ws_c.cell(row=r, column=1, value=f"  {a['name']}").font = df
        ws_c.cell(row=r, column=2, value=fmt(a["size"])).font = df
        ws_c.cell(row=r, column=2).alignment = Alignment(horizontal='right')
        ws_c.cell(row=r, column=3, value=method).font = df
        for ci in range(1, 4):
            ws_c.cell(row=r, column=ci).border = bd
        total_app += a["size"]
        r += 1

    ws_c.cell(row=r, column=1, value="  合计").font = Font(name="微软雅黑", size=10, bold=True)
    ws_c.cell(row=r, column=2, value=fmt(total_app)).font = Font(name="微软雅黑", size=10, bold=True, color="ED7D31")
    for ci in range(1, 4):
        ws_c.cell(row=r, column=ci).border = bd

    # 总计
    r += 2
    ws_c.cell(row=r, column=1, value="总计可释放").font = Font(name="微软雅黑", size=11, bold=True, color="C00000")
    ws_c.cell(row=r, column=2, value=fmt(total_safe + total_app)).font = Font(name="微软雅黑", size=11, bold=True, color="C00000")

    # 三、注意事项
    r += 2
    ws_c.cell(row=r, column=1, value="三、注意事项").font = stf
    notes = [
        "C:\\Windows、C:\\ProgramData、C:\\Program Files 不要手动删",
        "pagefile.sys 系统自动管理，不要动",
        "剪映/美图等缓存清理需要打开软件操作",
        "npm/pip 缓存可直接命令行清理",
        "建议每月跑一次 cleanmgr 保持磁盘整洁",
    ]
    for i, note in enumerate(notes):
        ws_c.cell(row=r+1+i, column=1, value=f"  {note}").font = df

    wb.save(path)
    return path


def main():
    print("=" * 60)
    print("  C 盘空间分析工具 v1.0")
    print("  正在扫描，请稍候...")
    print("=" * 60)
    if not is_admin():
        print("\n建议以管理员身份运行以获得完整结果\n")

    data = scan()

    desktop = os.path.join(os.path.expanduser("~"), "Desktop")
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    xp = os.path.join(desktop, f"C盘分析报告_{ts}.xlsx")

    try:
        gen_excel(data, xp)
        print(f"\nExcel 报告已生成: {xp}")
    except Exception as e:
        print(f"\n生成 Excel 失败: {e}")

    # 处理双击运行无 stdin 的情况
    def safe_input(prompt):
        try:
            return input(prompt).strip().lower()
        except (RuntimeError, EOFError):
            return "n"

    print("\n是否清理以下缓存?")
    r = safe_input("  清理 npm 缓存? (y/n): ")
    if r == "y":
        subprocess.run(["npm", "cache", "clean", "--force"],
                       capture_output=True, timeout=30)
        print("  npm 缓存已清理")
    r = safe_input("  清理 pip 缓存? (y/n): ")
    if r == "y":
        subprocess.run([sys.executable, "-m", "pip", "cache", "purge"],
                       capture_output=True, timeout=30)
        print("  pip 缓存已清理")

    print("\n完成!")
    try:
        input("按 Enter 退出...")
    except (RuntimeError, EOFError):
        pass


if __name__ == "__main__":
    main()
